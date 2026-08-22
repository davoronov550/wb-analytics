"""Export tests (T088) — writers + ExportProducts use case, offline (no DB)."""

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from analytics.adapters.outbound.export.writers import build_xlsx, iter_csv
from analytics.application.use_cases.export_products import ExportProducts
from catalog.application.dto import Ordering, Page, ProductFilter, ProductView


def _row(wb_id=1):
    return {
        "wb_id": wb_id,
        "name": "Наушники",
        "price": "100.00",
        "sale_price": "60.00",
        "discount_abs": "40.00",
        "discount_pct": "40.00",
        "rating": "4.5",
        "reviews_count": 100,
        "query": "наушники",
    }


class TestCsv:
    def test_header_then_rows(self):
        out = "".join(iter_csv([_row()]))
        lines = out.strip().splitlines()
        assert lines[0].startswith("wb_id")
        assert "Наушники" in lines[1]
        assert "60.00" in lines[1]

    def test_streams_line_by_line(self):
        chunks = list(iter_csv([_row(1), _row(2), _row(3)]))
        assert len(chunks) == 4  # header + 3 rows


class TestXlsx:
    def test_builds_valid_workbook_with_header_and_rows(self):
        data = build_xlsx([_row(7)])
        sheet = load_workbook(BytesIO(data)).active
        assert sheet.cell(row=1, column=1).value == "wb_id"
        assert sheet.cell(row=2, column=1).value == 7
        assert sheet.cell(row=2, column=2).value == "Наушники"


class FakeListProducts:
    def __init__(self, page):
        self.page = page
        self.calls = []

    def execute(self, filter, ordering, page, page_size):
        self.calls.append((filter, ordering, page, page_size))
        return self.page


def test_export_maps_views_and_passes_filter():
    view = ProductView(
        wb_id=1,
        name="A",
        price=Decimal("100.00"),
        sale_price=Decimal("60.00"),
        discount_abs=Decimal("40.00"),
        discount_pct=Decimal("40.00"),
        rating=Decimal("4.5"),
        reviews_count=100,
        query="q",
    )
    fake = FakeListProducts(Page(items=[view], count=1, page=1, page_size=100_000))

    rows = ExportProducts(list_products=fake).execute(
        ProductFilter(min_price=Decimal("50")), Ordering(field="price")
    )

    assert rows[0]["wb_id"] == 1
    assert rows[0]["sale_price"] == "60.00"
    assert fake.calls[0][0].min_price == Decimal("50")
