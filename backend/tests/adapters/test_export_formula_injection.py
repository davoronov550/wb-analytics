"""Exported cells must not execute as spreadsheet formulas.

Product names come from Wildberries — untrusted external input. A name starting
with =, +, -, @ (or a tab/CR) is executed by Excel and LibreOffice when the file
is opened, which is a data-exfiltration vector.
"""

import csv
import io

import pytest

from analytics.adapters.outbound.export.writers import build_xlsx, iter_csv

TRIGGERS = ["=", "+", "-", "@", "\t", "\r"]


def row(name: str) -> dict:
    return {
        "wb_id": 1,
        "name": name,
        "price": "10.00",
        "sale_price": "9.00",
        "discount_abs": "1.00",
        "rating": "5.0",
        "reviews_count": 1,
        "query": "q",
    }


def csv_cell(name: str) -> str:
    out = "".join(iter_csv(iter([row(name)])))
    return list(csv.reader(io.StringIO(out)))[1][1]


@pytest.mark.parametrize("trigger", TRIGGERS)
def test_formula_triggers_are_neutralised_in_csv(trigger):
    cell = csv_cell(f'{trigger}HYPERLINK("http://evil.test")')
    assert cell[:1] not in TRIGGERS, f"cell still starts with {cell[:1]!r}"


def test_ordinary_names_are_untouched_in_csv():
    assert csv_cell("Наушники A.Pods PRO 2") == "Наушники A.Pods PRO 2"


def test_numeric_columns_are_untouched():
    out = "".join(iter_csv(iter([row("ok")])))
    parsed = list(csv.reader(io.StringIO(out)))[1]
    assert parsed[2] == "10.00" and parsed[6] == "1"


@pytest.mark.parametrize("trigger", TRIGGERS)
def test_formula_triggers_are_neutralised_in_xlsx(trigger):
    from openpyxl import load_workbook

    payload = f"{trigger}cmd|' /c calc'!A0"
    workbook = load_workbook(io.BytesIO(build_xlsx(iter([row(payload)]))))
    cell = workbook["products"].cell(row=2, column=2).value
    assert str(cell)[:1] not in TRIGGERS
